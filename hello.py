import eel
from tkinter import filedialog
from tkinter import *
from openpyxl import load_workbook
import json, requests, ipaddress
requests.packages.urllib3.disable_warnings()

# Set web files folder and optionally specify which file types to check for eel.expose()
#   *Default allowed_extensions are: ['.js', '.html', '.txt', '.htm', '.xhtml']
eel.init('web', allowed_extensions=['.js', '.html'])

def sendupdate(return_html):
    eel.my_javascript_function(return_html)

### Start copy from draft

def openbook(filename):
    try:
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb['Sheet1']

        ## Get Columns
        headings = ['nul']
        col = 0
        blankheading = 0
        while blankheading < 3:
            col += 1
            if ws.cell(row=1, column=col).value == None:
                blankheading += 1
            else:
                headings.append(ws.cell(row=1, column=col).value)

        ## Get all Device Rows
        AllDevicesList = []
        device_meta_data = {}
        device_dint_data = {}
        device_sdwanint_data = {}
        device_daddr_data = {}
        blankrow = 0
        row = 1

        while blankrow < 3:
            row += 1
            if ws.cell(row=row, column=1).value is None:
                blankrow += 1
            else:
                # get device detail in row

                col = 1
                newdict = {}
                for i in headings:
                    if i != 'nul':
                        newdict[i] = ws.cell(row=row, column=col).value
                        if i == "Device_name":
                            device_meta_data[newdict['Device_name']] = {}
                            device_dint_data[newdict['Device_name']] = {}
                            device_sdwanint_data[newdict['Device_name']] = {}
                            device_daddr_data[newdict['Device_name']] = {}
                        if i[0:5] == "meta_":
                            device_meta_data[newdict['Device_name']][i[5:]] = ws.cell(row=row, column=col).value
                        if i[0:5] == "dint_":
                            device_dint_data[newdict['Device_name']][i[5:]] = ws.cell(row=row, column=col).value
                        if i[0:9] == "sdwanint_":
                            sdwanintsettings = i[9:].split("|")
                            try:
                                device_sdwanint_data[newdict['Device_name']][sdwanintsettings[0]]
                            except:
                                device_sdwanint_data[newdict['Device_name']][sdwanintsettings[0]] = {}
                            device_sdwanint_data[newdict['Device_name']][sdwanintsettings[0]][sdwanintsettings[1]] = ws.cell(row=row, column=col).value

                        if i[0:6] == "daddr_":
                            device_daddr_data[newdict['Device_name']][i[6:]] = ws.cell(row=row, column=col).value



                        col += 1

                AllDevicesList.append(newdict)
    except:
        AllDevicesList = "failed"

    try:

        ## build supported devices

        ws2 = wb['Supported Devices']
        blankrow = 0
        device_platform_data = {}
        row = 1

        while blankrow < 3:
            row += 1
            if ws2.cell(row=row, column=1).value == None:
                blankrow += 1
            else:
                device_platform_data[ws2.cell(row=row, column=1).value] = {
                    "platform": ws2.cell(row=row, column=2).value,
                    "platform_id": ws2.cell(row=row, column=3).value,
                    "os_ver": ws2.cell(row=row, column=7).value,
                    "version": ws2.cell(row=row, column=9).value,
                    "mr": ws2.cell(row=row, column=8).value,
                    "prefer_img_ver": ws2.cell(row=row, column=6).value,
                    "branch_pt": ws2.cell(row=row, column=5).value,
                    "build": ws2.cell(row=row, column=4).value

                }
    except:
        device_platform_data = "failed"

    return AllDevicesList, device_platform_data, headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data


def get_meta():
    requestid = 2
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/dvmdb/_meta_fields/device",
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    meta_json = json.loads(res.text)
    return meta_json['result'][0]['data']


def create_meta(newname):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/dvmdb/_meta_fields/device",
                "data": {
                    "importance": 0,
                    "length": 255,
                    "name": newname,
                    "status": 1
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    try:
        metacreate_data = json.loads(res.text)
        ret_meta = "Meta Field " + newname + " created.<br>\n"
    except:
        ret_meta = "Failed to create Meta Field " + newname
    return ret_meta

def track_model_task(taskid):

    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)

    return ret_status


def add_model_device(adomname,devicename,sn,devicetype):

    version = device_platform_data[devicetype]['version']
    os_ver = device_platform_data[devicetype]['os_ver']
    major_rev = device_platform_data[devicetype]['mr']
    platform = device_platform_data[devicetype]['platform']
    prefer_img = device_platform_data[devicetype]['prefer_img_ver']
    platform_id = device_platform_data[devicetype]['platform_id']
    branch = device_platform_data[devicetype]['branch_pt']
    build = device_platform_data[devicetype]['build']

    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvm/cmd/add/device",
                "data": {
                    "adom": adomname,
                    "flags": [
                        "create_task",
                        "nonblocking"
                    ],
                    "device": {
                        "name": devicename,
                        "os_ver": os_ver,
                        "adm_usr": "admin",
                        "version": version,
                        "os_type": 0,
                        "mr": major_rev,
                        "platform_str": platform,
                        "prefer_img_ver": prefer_img,
                        "mgmt_mode": 3,
                        "flags": 67371040,
                        "sn": sn,
                        "faz.perm": 15,
                        "faz.quota": 0,
                        "platform_id": platform_id,
                        "branch_pt": branch,
                        "build": build
                    }
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['taskid'])
    return last_task

def update_device(adom,devicename):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "meta fields": device_meta_data[devicename],
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)

def assign_cli_template(adom,template,devicename):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/template-group",
                "data": {
                    "name": template,
                    "scope member": [
                        {
                            "name": devicename,
                            "vdom": "root"
                        }
                    ]
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(res.text)
    json_assignclitemplate = json.loads(res.text)
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def quickinstall(adom,devicename,vdom):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "/securityconsole/install/device",
                "data": {
                    "adom": adom,
                    "scope": [
                        {
                            "name": devicename,
                            "vdom": vdom
                        }
                    ]
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_quickinstall = json.loads(res.text)
    taskid_qi = json_quickinstall['result'][0]['data']['task']
    return taskid_qi

def track_quickinstall(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)
    return ret_status

def add_install_target(device,adomname,vdomname,pkg):
    requestid = 1
    jsondata = {
      "method":"add",
      "params":[
        {
          "url": "pm/pkg/adom/" + adomname + "/" + pkg + "/scope member",
          "data": [
            {
                "name": device,
                "vdom": vdomname
            },
          ]
        }
      ],
      "id": requestid,
      "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(res.text)
    json_assignppkg = json.loads(res.text)
    status_ppkg = json_assignppkg['result'][0]['status']['message']
    return status_ppkg


def install_pkg(pkg,adomname,devicename,vdom):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "securityconsole/install/package",
                "data":
                    { "pkg": pkg,
                      "adom": adomname,
                      "scope": [
                          {
                              "name": devicename,
                              "vdom": vdom
                          }
                      ]
                    }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print()
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['task'])
    return last_task



def track_policyinstall(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)
    return ret_status

def add_policy_interface_member(adomname, newinterfacename, realinterface, devicename):
    requestid = 1
    jsondata = {
      "method":"add",
      "params":[
        {
          "url":"pm/config/adom/" + adomname + "/obj/dynamic/interface/" + newinterfacename + "/dynamic_mapping",
          "data":
                  {
                      "_scope": [
                          {
                              "name": devicename,
                              "vdom": "root"
                          }
                      ],
                      "local-intf": [
                          realinterface
                      ],
                      "intrazone-deny": 0
                  }

        }
      ],
      "id":requestid,
      "session":fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_mapdint = json.loads(res.text)
    status_mapdint = json_mapdint['result'][0]['status']['message']
    return status_mapdint


def add_sdwaninterface_mapping(adomname,devicename,interfacename,vdom):
    ## get settings for base SDWAN interface
    requestid = 1

    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/" + interfacename,
                }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_sdwanint_res = json.loads(res.text)
    json_sdwanint= json_sdwanint_res['result'][0]['data']
    json_sdwanint.pop('dynamic_mapping', None)
    json_sdwanint.pop('obj seq', None)
    json_sdwanint.pop('name', None)

    json_sdwanint["_scope"] = [
        {
            "name": devicename,
            "vdom": vdom
        }
    ]

    for key in device_sdwanint_data[devicename][interfacename]:
        json_sdwanint[key] = device_sdwanint_data[devicename][interfacename][key]

    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/" + interfacename + "/dynamic_mapping",
                "data": json_sdwanint
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(json_sdwanint)
    print(res.text)
    json_mapsdwanint = json.loads(res.text)
    status_mapsdwanint = json_mapsdwanint['result'][0]['status']['message']
    return status_mapsdwanint

def assign_sdwan_template(adom,sdwantemplate,devicename,vdom):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/pm/wanprof/adom/" + adom + "/" + sdwantemplate + "/scope member",
                "data": [
                        {
                            "name": devicename,
                            "vdom": vdom
                        }
                    ]
                }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(res.text)
    json_assignclitemplate = json.loads(res.text)
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate

def add_daddr(adomname,daddrobj,newaddr,devicename,vdom):
    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/firewall/address/" + daddrobj
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(res.text)

    current_int_result = json.loads(res.text)
    current_int = current_int_result['result'][0]['data']

    result_msg = "unknown error"

    submit = False

    if current_int['type'] is 0:
        try:
            addrsettings = [
                {
                    "_scope": [
                        {
                            "name": devicename,
                            "vdom": vdom
                        }
                    ],

                    "subnet": [
                        str(ipaddress.ip_network(newaddr).network_address),
                        str(ipaddress.ip_network(newaddr).netmask)
                    ],
                }
            ]
            submit = True
        except:
            result_msg = "WARNING: Could not decode ip address into network_address/netmask"
    elif current_int['type'] is 1:
        try:
            newaddr.strip(" ")
            splitaddr = newaddr.split("-")


            addrsettings = [
                {
                    "_scope": [
                        {
                            "name": devicename,
                            "vdom": vdom
                        }
                    ],

                    "end-ip": splitaddr[1].strip(),
                    "start-ip": splitaddr[0].strip()
                }
            ]
            submit = True
        except:
            result_msg = "WARNING: Could not calculate IP RANGE"

    if submit is True:

        requestid = 1
        jsondata = {
            "method": "add",
            "params": [
                {
                    "url": "pm/config/adom/" + adomname + "/obj/firewall/address/" + daddrobj + "/dynamic_mapping",
                    "data": addrsettings
                }
            ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        print(res.text)
        json_result = json.loads(res.text)
        result_msg = json_result['result'][0]['status']['message']
    return result_msg




@eel.expose
def btn_checkxlsx(filename,fmghost,fmguser,fmgpasswd,fmgadom):
    global fmg_user
    global fmg_passwd
    global fmgurl
    global fmg_adom
    global fmg_sessionid
    global device_platform_data
    global device_meta_data
    global device_dint_data
    global device_sdwanint_data
    global alldevices

    qi_status = False
    fmg_adom = fmgadom
    fmg_user = fmguser
    fmg_passwd = fmgpasswd
    fmgurl = "https://" + fmghost + "/jsonrpc"
    fmg_sessionid = None

    alldevices, device_platform_data, headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data = openbook(filename)
    return_html = ""
    if alldevices == "failed" or device_platform_data == "failed":
        return_html += "Load Excel workbook failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
    else:
        return_html += "Load Excel workbook successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
        ##login to FMG

        requestid = 1

        jsondata = {'method': 'exec',
                    'params': [{'url': '/sys/login/user', 'data': {'user': fmg_user, 'passwd': fmg_passwd}}],
                    'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)
        try:
            login_data = json.loads(res.text)
            fmg_sessionid = login_data['session']
            return_html += "FortiManager login successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
        except:
            return_html += "FortiManager login failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    if fmg_sessionid is not None:
        proceed = True
    else:
        proceed = False


    ### validity checks
    if proceed == True:
        ## Does ADOM exist in FMG

        jsondata = {"method": "get", "params": [{"url": "dvmdb/adom/" + fmg_adom}], "id": requestid,
                    "session": fmg_sessionid}
        res = session.post(fmgurl, json=jsondata, verify=False)
        json_result = json.loads(res.text)
        print (json_result['result'][0]['status']['message'])
        if json_result['result'][0]['status']['message'] != "OK":
            return_html += "FortiManager ADOM does not exist <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
            proceed = False

        ## check for required fields in headings
        required_headings = ["Device_name", "Device_Type", "Device_SN", "CLI_Template", "Policy_Package", "SDWAN_Template"]
        for req_heading in required_headings:
            if req_heading not in headings:
                return_html += "Excel File does not have required heading \"" + req_heading + "\" <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                proceed = False


    ### Check for Meta Fields and Create if they dont exist
    if proceed == True:

        metafields = get_meta()
        for field in headings:
            if field[0:5] == "meta_":
                try:
                    IPDICT = next(item for item in metafields if item["name"] == field[5:])
                except:
                    create_meta(field[5:])


    ### Create Model Devices
    if proceed == True:
        sendupdate(return_html)

        for devicedata in alldevices:
            return_html += "<br>\n <b> >> Adding Device [ " + devicedata['Device_name'] + " ] </b><br>\n"
            add_dev_status = track_model_task(add_model_device(fmg_adom, devicedata['Device_name'], devicedata['Device_SN'],
                                        devicedata['Device_Type']))

            sendupdate(return_html)

            if add_dev_status == True:
                return_html += "Adding model device (" + devicedata['Device_name'] + "/" + devicedata['Device_SN'] + ") successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
            else:
                return_html += "Adding model device (" + devicedata['Device_name'] + "/" + devicedata['Device_SN'] + ") failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

            if add_dev_status == True:
                ## Add meta data to device
                update_device(fmg_adom, devicedata['Device_name'])

                ## Assign CLI Template
                status_clitemp = assign_cli_template(fmg_adom, devicedata['CLI_Template'], devicedata['Device_name'])
                if status_clitemp == "OK":
                    return_html += "Assign CLI template Group \"" + devicedata['CLI_Template'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                elif status_clitemp == "Object does not exist":
                    return_html += "Assign CLI template Group \"" + devicedata['CLI_Template'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                    return_html += "CLI Template not found <br>\n"
                else:
                    return_html += "Assign CLI template Group \"" + devicedata['CLI_Template'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                if status_clitemp == "OK":
                    ##Install Device Settings
                    qi_status = track_quickinstall(quickinstall(fmg_adom,devicedata['Device_name'],'root'))
                    if qi_status:
                        return_html += "Quick install device settings successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                    else:
                        return_html += "Quick install device settings failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

            if qi_status == True:
                ## Quick Install Sucessful, assign policy package etc
                sendupdate(return_html)
                ##map interfaces

                for key in device_dint_data[devicedata['Device_name']]:
                    status_mapdint = add_policy_interface_member(fmg_adom, key,
                                                                 device_dint_data[devicedata['Device_name']][key],
                                                                 devicedata['Device_name'])

                    if status_mapdint == "OK":
                        return_html += "Add dynamic map for interface \"" + key + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                    else:
                        return_html += "Add dynamic map for interface \"" + key + "\" failed ><span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                        return_html += status_mapdint + "<br>\n"

                ### MAP DYNAMIC ADDRESS OJBECTS

                for key in device_daddr_data[devicedata['Device_name']]:
                    status_mapdaddr = add_daddr(fmg_adom,key,device_daddr_data[devicedata['Device_name']][key],devicedata['Device_name'],'root')
                    if status_mapdaddr == "OK":
                        return_html += "Add dynamic map for address \"" + key + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                    else:
                        return_html += "Add dynamic map for address \"" + key + "\" failed ><span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                        return_html += status_mapdaddr + "<br>\n"




                ### MAP SDWAN Interfaces

                for key in device_sdwanint_data[devicedata['Device_name']]:
                    statu_mapsdwanint = add_sdwaninterface_mapping(fmg_adom, devicedata['Device_name'], key, 'root')
                    if statu_mapsdwanint == "OK":
                        return_html += "Add dynamic SDWAN Map for interface \"" + key + "\" succcessful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                    else:
                        return_html += "Add dynamic SDWAN Map for interface \"" + key + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                ## Assign SDWAN Template

                status_assignsdwantemplate = assign_sdwan_template(fmg_adom, devicedata['SDWAN_Template'], devicedata['Device_name'], 'root')
                if status_assignsdwantemplate == "OK":
                    return_html += "Assign SDWAN template \"" + devicedata['SDWAN_Template'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                else:
                    return_html += "Assign SDWAN template \"" + devicedata['SDWAN_Template'] + "\" failed ><span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"




                ## Add install Target
                status_add_inst_trgt = add_install_target(devicedata['Device_name'], fmg_adom, "root", devicedata['Policy_Package'])
                if status_add_inst_trgt == "OK":
                    return_html += "Assign policy package \"" + devicedata['Policy_Package'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                else:
                    return_html += "Assign policy package \"" + devicedata['Policy_Package'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                if status_add_inst_trgt == "OK":
                    ## install package
                    pkg_status = track_policyinstall(install_pkg(devicedata['Policy_Package'], fmg_adom, devicedata['Device_name'], 'root'))
                    if pkg_status == True:
                        return_html += "Install policy package \"" + devicedata['Policy_Package'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                    else:
                        return_html += "Install policy package \"" + devicedata['Policy_Package'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    if pkg_status == True:
                        pass

            sendupdate(return_html)


    ### LOGOUT OF FMG
    if fmg_sessionid is not None:
        requestid = 1
        jsondata = {'method': 'exec', 'params': [{'url': '/sys/logout'}], 'session': fmg_sessionid, 'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)

    return_html += "<br>\n<b> >> Complete! <br>\n"
    return_html += "<br>\n <a href=\"hello.html\">Return</a> <br>\n"

    sendupdate(return_html)

### End copy from draft

@eel.expose
def btn_checkadom(filename,fmghost,fmguser,fmgpasswd,fmgadom,fmgadomdesc):
    global fmg_user
    global fmg_passwd
    global fmgurl
    global fmg_adom
    global fmg_adomdesc
    global fmg_sessionid

    fmg_adom = fmgadom
    fmg_user = fmguser
    fmg_passwd = fmgpasswd
    fmgurl = "https://" + fmghost + "/jsonrpc"
    fmg_sessionid = None

    return_html = ""

    new_json = ""
    with open(filename) as json_data_file:
        vars = {"$(adom_name)": fmgadom, "$(adom_desc)": fmgadomdesc}

        for line in json_data_file:
            m = re.search('\$\((.+)\)', line)
            if m:
                print("Old: " + line.rstrip())
                for key in vars.keys():
                    line = line.replace(key, vars[key])

                print("New: " + line.rstrip())

            new_json += line

    try:
        adom_json = json.loads(new_json)
        return_html += "Load JSON file successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
    except:
        return_html += "Load Excel workbook failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    sendupdate(return_html)

    if adom_json:
        ##login to FMG

        requestid = 1

        jsondata = {'method': 'exec',
                    'params': [{'url': '/sys/login/user', 'data': {'user': fmg_user, 'passwd': fmg_passwd}}],
                    'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)
        try:
            login_data = json.loads(res.text)
            fmg_sessionid = login_data['session']
            return_html += "FortiManager login successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
        except:
            return_html += "FortiManager login failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    sendupdate(return_html)

    if fmg_sessionid is not None:
        for item in adom_json["settings"]:
            for dataset in item["data"]:

                jsondata = {
                    "method": item["method"],
                    "params": [
                        {
                            "url": item["url"],
                            "data": dataset

                        }
                    ],
                    "id": requestid,
                    "session": fmg_sessionid
                }
                print("Request:")
                print(json.dumps(jsondata, indent=4, sort_keys=True))
                res = session.post(fmgurl, json=jsondata, verify=False)
                response = json.loads(res.text)
                print("Response:")
                print(json.dumps(response, indent=4, sort_keys=True))
                response_url = response['result'][0]['url']
                try:
                    response_name = response['result'][0]['data']['name']
                except:
                    try:
                        response_name = jsondata['params'][0]['data']['name']
                    except:
                        response_name = ""
                if response['result'][0]['status']['message'] == "OK":
                    return_html += response_url + response_name + " <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                else:
                    return_html += response_url + response_name + " <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                    return_html += " >> " + response['result'][0]['status']['message'] + "<br>\n"

                sendupdate(return_html)

    ### LOGOUT OF FMG
    if fmg_sessionid is not None:
        requestid = 1
        jsondata = {'method': 'exec', 'params': [{'url': '/sys/logout'}], 'session': fmg_sessionid, 'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)

    return_html += "<br>\n<b> >> Complete! <br>\n"
    return_html += "<br>\n <a href=\"hello.html\">Return</a> <br>\n"

    sendupdate(return_html)


@eel.expose
def btn_ResimyoluClick():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("XLSX Files", "*.xlsx"), ("all files", "*.*")))


    print(filename)
    return filename

@eel.expose
def btn_getjsonfile():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("JSON Files", "*.json"), ("all files", "*.*")))

    print(filename)
    return filename



@eel.expose
def getsettings_adom():
    try:
        with open('settings.json') as json_settings:
            settings = json.load(json_settings)
            default_fmg = settings['fmg']
            default_user = settings['user']
    except:
        default_fmg = ""
        default_user = ""

    return_html = '''
            <div class="starter-template">
      <h3>Import ADOM</h3>
    </div>
    <div>
      <form autocomplete="off">

        <label for="fmgip">FortiManager URL</label>



        <div class="input-group mb-3">
          <div class="input-group-prepend">
            <span class="input-group-text" id="https-addon">https://</span>
          </div>
          <input type="text" class="form-control" id="fmgip" aria-describedby="https-addon" value="%s">
        </div>


        <div class="form-group">
          <label for="fmgusername">FortiManager Username</label>
          <input type="text" class="form-control" id="fmgusername" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgpassword">FortiManager Password</label>
          <input type="password" class="form-control" id="fmgpassword">
        </div>
        <div class="form-group">
          <label for="fmgadom">New FortiManager ADOM</label>
          <input type="text" class="form-control" id="fmgadom" value="">
        </div>
        <div class="form-group">
          <label for="fmgadom">New FortiManager ADOM Description</label>
          <input type="text" class="form-control" id="fmgadomdesc" value="">
        </div>        
        <div form-group>
          <button type="button" onclick="getFileADOM()" class="btn btn-secondary">Select File</button>
          Excel Path: <span id="filepath">/</span><br/><br/>
        </div>
        <div class="form-group">
          <button type="button" onclick="processadom(document.getElementById('filepath').innerHTML)" class="btn btn-primary">Submit</button>
        </div>
      </form>

    </div>
          ''' % (default_fmg, default_user)

    eel.my_javascript_function(return_html)

@eel.expose
def getsettings_devices():
    try:
        with open('settings.json') as json_settings:
            settings = json.load(json_settings)
            default_fmg = settings['fmg']
            default_user = settings['user']
    except:
        default_fmg = ""
        default_user = ""

    return_html = '''
            <div class="starter-template">
      <h3>ZTP Tool</h3>
      <p class="lead">Load Excel file to create model devices in FortiManager.</p>
    </div>
    <div>
      <form autocomplete="off">

        <label for="fmgip">FortiManager URL</label>



        <div class="input-group mb-3">
          <div class="input-group-prepend">
            <span class="input-group-text" id="https-addon">https://</span>
          </div>
          <input type="text" class="form-control" id="fmgip" aria-describedby="https-addon" value="%s">
        </div>


        <div class="form-group">
          <label for="fmgusername">FortiManager Username</label>
          <input type="text" class="form-control" id="fmgusername" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgpassword">FortiManager Password</label>
          <input type="password" class="form-control" id="fmgpassword">
        </div>
        <div class="form-group">
          <label for="fmgadom">FortiManager ADOM</label>
          <input type="text" class="form-control" id="fmgadom" value="">
        </div>
        <div form-group>
          <button type="button" onclick="getFolder()" class="btn btn-secondary">Select File</button>
          Excel Path: <span id="filepath">/</span><br/><br/>
        </div>
        <div class="form-group">
          <button type="button" onclick="processxlsx(document.getElementById('filepath').innerHTML)" class="btn btn-primary">Submit</button>
        </div>
      </form>

    </div>
          ''' % (default_fmg, default_user)

    eel.my_javascript_function(return_html)



session = requests.session()

eel.start('hello.html', size=(900, 900), disable_cache=True)             # Start (this blocks and enters loop)