import time
import eel
from tkinter import filedialog
from tkinter import *
from tkinter import *
from openpyxl import load_workbook
import json, requests, ipaddress, sys, platform, io

requests.packages.urllib3.disable_warnings()

# Set web files folder and optionally specify which file types to check for eel.expose()
#   *Default allowed_extensions are: ['.js', '.html', '.txt', '.htm', '.xhtml']
eel.init('web', allowed_extensions=['.js', '.html'])


def sendupdate(return_html):
    eel.pageupdate(return_html)


### Export ADOM Functions

def export_adom(adomname):
    global export_info
    global sdwan_template_list
    global polpkg_list
    export_info = {"vars": ["adom_name", "adom_desc"], "settings": []}
    requestid = 1

    ## Get ADOM Info
    newdict = {"url": "/dvmdb/adom/", "method": "add", "data": []}

    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/dvmdb/adom/" + adomname
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    parsed = json.loads(res.text)
    # print(json.dumps(parsed, indent=4, sort_keys=True))

    dvmdbadom = parsed['result'][0]['data']
    dvmdbadom.pop('uuid', None)
    dvmdbadom.pop('oid', None)
    dvmdbadom["name"] = "$(adom_name)"
    dvmdbadom["desc"] = "$(adom_desc)"
    if adomname == "root":
        dvmdbadom["flags"] = 2312

    newdict["data"].append(dvmdbadom)
    export_info["settings"].append(newdict)

    ## standard objects (objects which can be exported and imported with out changing anything, mutliple objects can be created in a list)

    sdwan_template_list = []
    polpkg_list = []

    std_objects = {
        "clitemplate": ["/pm/config/adom/$(adom_name)/obj/cli/template",
                        "/pm/config/adom/" + adomname + "/obj/cli/template", ["modification-time"]],
        "clitemplate-group": ["/pm/config/adom/$(adom_name)/obj/cli/template-group",
                              "/pm/config/adom/" + adomname + "/obj/cli/template-group", ["modification-time"]],
        "sdwaninterface": ["/pm/config/adom/$(adom_name)/obj/dynamic/virtual-wan-link/members/",
                           "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/",
                           ["dynamic_mapping", "obj seq"]],
        "sdwanservers": ["/pm/config/adom/$(adom_name)/obj/dynamic/virtual-wan-link/server/",
                         "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/server/",
                         ["dynamic_mapping"]],
        "sdwantemplates": ["/pm/wanprof/adom/$(adom_name)",
                           "/pm/wanprof/adom/" + adomname,
                           ["scope member", "oid"]],
        "addrobjs": ["pm/config/adom/$(adom_name)/obj/firewall/address/",
                     "pm/config/adom/" + adomname + "/obj/firewall/address/",
                     ["uuid", "dynamic_mapping"]],
        "addrobjs_grp": ["pm/config/adom/$(adom_name)/obj/firewall/addrgrp/",
                         "pm/config/adom/" + adomname + "/obj/firewall/addrgrp/",
                         ["uuid", "dynamic_mapping"]],
        "intfobjs": ["pm/config/adom/$(adom_name)/obj/dynamic/interface/",
                     "pm/config/adom/" + adomname + "/obj/dynamic/interface/",
                     ["uuid", "dynamic_mapping"]],
        "applist": ["pm/config/adom/$(adom_name)/obj/application/list",
                    "pm/config/adom/" + adomname + "/obj/application/list",
                    ["uuid", "dynamic_mapping", "obj seq"]],
        "appgrp": ["pm/config/adom/$(adom_name)/obj/application/group",
                   "pm/config/adom/" + adomname + "/obj/application/group",
                   ["uuid", "dynamic_mapping", "obj seq"]],
        "service": ["pm/config/adom/$(adom_name)/obj/firewall/service/custom",
                    "pm/config/adom/" + adomname + "/obj/firewall/service/custom",
                    ["uuid", "dynamic_mapping", "obj seq"]],
        "servicegrp": ["pm/config/adom/$(adom_name)/obj/firewall/service/group",
                       "pm/config/adom/" + adomname + "/obj/firewall/service/group",
                       ["uuid", "dynamic_mapping", "obj seq"]],
        "polpkg": ["pm/pkg/adom/$(adom_name)",
                   "pm/pkg/adom/" + adomname,
                   ["scope member", "oid"]]

    }

    for objecturls in std_objects:
        get_and_add(std_objects, objecturls)

    ## Get SDWAN Template Details
    for sdwan_template in sdwan_template_list:
        get_and_add({"sdwan_member": [
            "pm/config/adom/$(adom_name)/wanprof/" + sdwan_template + "/system/virtual-wan-link/member",
            "pm/config/adom/" + adomname + "/wanprof/" + sdwan_template + "/system/virtual-wan-link/member",
            ["obj seq"]]}, "sdwan_member")

        get_and_add({"sdwan_hlth": [
            "pm/config/adom/$(adom_name)/wanprof/" + sdwan_template + "/system/virtual-wan-link/health-check",
            "pm/config/adom/" + adomname + "/wanprof/" + sdwan_template + "/system/virtual-wan-link/health-check",
            ["obj seq"]]}, "sdwan_hlth")

        get_and_add({"sdwan_service": [
            "pm/config/adom/$(adom_name)/wanprof/" + sdwan_template + "/system/virtual-wan-link/service",
            "pm/config/adom/" + adomname + "/wanprof/" + sdwan_template + "/system/virtual-wan-link/service",
            ["obj seq"]]}, "sdwan_service")

    ## Get Policy Package Details
    for polpkg in polpkg_list:
        get_and_add({"polpkg_policy": ["pm/config/adom/$(adom_name)/pkg/" + polpkg + "/firewall/policy",
                                       "pm/config/adom/" + adomname + "/pkg/" + polpkg + "/firewall/policy",
                                       ["obj seq", "_policy_block"]]}, "polpkg_policy")

    # print(json.dumps(export_info, indent=4, sort_keys=True))
    return json.dumps(export_info, indent=4, sort_keys=True)


def get_and_add(std_objects, objecturls):
    newdict = {"url": std_objects[objecturls][0], "method": "add", "data": []}
    if objecturls == "sdw_members":
        newdict = {"url": std_objects[objecturls][0], "method": "replace", "data": []}
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": std_objects[objecturls][1]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    parsed = json.loads(res.text)
    # print(json.dumps(parsed, indent=4, sort_keys=True))

    newdata = parsed['result'][0]['data']
    for index, config in enumerate(newdata):
        for popitem in std_objects[objecturls][2]:
            if popitem in newdata[index].keys():
                newdata[index].pop(popitem)

        ignore_addr_obj = ["wildcard.dropbox.com", "wildcard.google.com", "SSLVPN_TUNNEL_ADDR1", "all", "gmail.com",
                           "login.microsoft.com", "login.microsoftonline.com", "login.windows.net", "none",
                           "FABRIC_DEVICE", "FIREWALL_AUTH_PORTAL_ADDRESS"]
        if objecturls == "addrobjs":
            if newdata[index]['name'] in ignore_addr_obj:
                newdata[index] = {}

        ignore_addrgrp_obj = ["G Suite", "Microsoft Office 365"]
        if objecturls == "addrobjs_grp":
            if newdata[index]['name'] in ignore_addrgrp_obj:
                newdata[index] = {}

        ignore_sdwan_hlth = ["Default_AWS", "Default_FortiGuard", "Default_Gmail", "Default_Google Search",
                             "Default_Office_365"]
        if objecturls == "sdwan_hlth":
            if newdata[index]['name'] in ignore_sdwan_hlth:
                newdata[index] = {}

        if objecturls == "sdwantemplates":
            sdwan_template_list.append(newdata[index]['name'])

        if objecturls == "polpkg":
            polpkg_list.append(newdata[index]['name'])

        if objecturls == "sdwan_service":
            if isinstance(newdata[index]['sla'], list):
                for index2, config2 in enumerate(newdata[index]['sla']):
                    if "obj seq" in newdata[index]['sla'][index2].keys():
                        newdata[index]['sla'][index2].pop("obj seq")

        ignore_applist_obj = ["block-high-risk", "default", "sniffer-profile", "wifi-default"]
        if objecturls == "applist":
            if isinstance(newdata[index]['entries'], list):
                for index2, config2 in enumerate(newdata[index]['entries']):
                    if "obj seq" in newdata[index]['entries'][index2].keys():
                        newdata[index]['entries'][index2].pop("obj seq")

            if newdata[index]['name'] in ignore_applist_obj:
                newdata[index] = {}

        ignore_service_obj = ["ALL", "ALL_TCP", "ALL_UDP", "ALL_ICMP", "ALL_ICMP6", "GRE", "GTP", "AH", "ESP", "AOL",
                              "BGP", "DHCP", "DNS", "FINGER", "FTP", "FTP_GET", "FTP_PUT", "GOPHER", "H323", "HTTP",
                              "HTTPS", "IKE", "IMAP", "IMAPS", "Internet-Locator-Service", "IRC", "L2TP", "LDAP",
                              "NetMeeting", "NFS", "NNTP", "NTP", "OSPF", "PC-Anywhere", "PING", "TIMESTAMP",
                              "INFO_REQUEST", "INFO_ADDRESS", "ONC-RPC", "DCE-RPC", "POP3", "POP3S", "PPTP", "QUAKE",
                              "RAUDIO", "REXEC", "RIP", "RLOGIN", "RSH", "SCCP", "SIP", "SIP-MSNmessenger", "SAMBA",
                              "SMTP", "SMTPS", "SNMP", "SSH", "SYSLOG", "TALK", "TELNET", "TFTP", "MGCP", "UUCP",
                              "VDOLIVE", "WAIS", "WINFRAME", "X-WINDOWS", "PING6", "MS-SQL", "MYSQL", "RDP", "VNC",
                              "DHCP6", "SQUID", "SOCKS", "WINS", "RADIUS", "RADIUS-OLD", "CVSPSERVER", "AFS3",
                              "TRACEROUTE", "RTSP", "MMS", "KERBEROS", "LDAP_UDP", "SMB", "NONE", "webproxy"]
        if objecturls == "service":
            if newdata[index]['name'] in ignore_service_obj:
                newdata[index] = {}

        ignore_servicegrp_obj = ["Email Access", "Web Access", "Windows AD", "Exchange Server"]
        if objecturls == "servicegrp":
            if newdata[index]['name'] in ignore_servicegrp_obj:
                newdata[index] = {}

    while {} in newdata:
        newdata.remove({})
    newdict["data"] = newdata
    export_info["settings"].append(newdict)


### Start copy from draft

def openbook(filename):
    headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data, device_daddr6_data, device_vpn_data = "", "", "", "", "", "", ""
    try:
        with open(filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        wb = load_workbook(in_mem_file, data_only=True)

        try:
            ws = wb['Devices']
        except:
            try:
                ws = wb['Sheet1']
            except:
                ws = wb.active

        print("cell A1 value = " + ws.cell(1, 1).value)

        if ws.cell(1, 1).value == "Device_Name":
            print("ok")
            ## Get Columns
            headings = ['nul']
            col = 0
            blankheading = 0
            while blankheading < 3:
                col += 1
                if ws.cell(row=1, column=col).value == None:
                    blankheading += 1
                else:
                    headings.append(ws.cell(row=1, column=col).value)

            ## Get all Device Rows
            AllDevicesList = []
            device_meta_data = {}
            device_dint_data = {}
            device_sdwanint_data = {}
            device_daddr_data = {}
            device_vpn_data = {}
            device_daddr6_data = {}
            blankrow = 0
            row = 1

            while blankrow < 3:
                row += 1
                if ws.cell(row=row, column=1).value is None:
                    blankrow += 1
                else:
                    # get device detail in row

                    col = 1
                    newdict = {}
                    for i in headings:
                        if i != 'nul':
                            if ws.cell(row=row, column=col).value is None:
                                newdict[i] = ""
                            else:
                                newdict[i] = str(ws.cell(row=row, column=col).value)
                            if i == "Device_Name":
                                device_meta_data[newdict['Device_Name']] = {}
                                device_meta_data[newdict['Device_Name']]['Device_Name'] = newdict['Device_Name']
                                device_dint_data[newdict['Device_Name']] = {}
                                device_sdwanint_data[newdict['Device_Name']] = {}
                                device_daddr_data[newdict['Device_Name']] = {}
                                device_vpn_data[newdict['Device_Name']] = {}
                                device_daddr6_data[newdict['Device_Name']] = {}
                            if i == "Device_SN":
                                device_meta_data[newdict['Device_Name']]['Device_SN'] = newdict['Device_SN']
                            if i[0:5] == "meta_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_meta_data[newdict['Device_Name']][i[5:]] = ""
                                else:
                                    device_meta_data[newdict['Device_Name']][i[5:]] = str(
                                        ws.cell(row=row, column=col).value)
                            if i[0:5] == "dint_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_dint_data[newdict['Device_Name']][i[5:]] = ""
                                else:
                                    device_dint_data[newdict['Device_Name']][i[5:]] = str(
                                        ws.cell(row=row, column=col).value).split(",")
                            if i[0:9] == "sdwanint_":
                                sdwanintsettings = i[9:].split("|")
                                try:
                                    device_sdwanint_data[newdict['Device_Name']][sdwanintsettings[0]]
                                except:
                                    device_sdwanint_data[newdict['Device_Name']][sdwanintsettings[0]] = {}
                                if ws.cell(row=row, column=col).value is not None:
                                    device_sdwanint_data[newdict['Device_Name']][sdwanintsettings[0]][
                                        sdwanintsettings[1]] = str(ws.cell(row=row, column=col).value)

                            if i[0:6] == "daddr_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_daddr_data[newdict['Device_Name']][i[6:]] = ""
                                else:
                                    device_daddr_data[newdict['Device_Name']][i[6:]] = str(
                                        ws.cell(row=row, column=col).value)
                            if i[0:7] == "daddr6_":
                                if ws.cell(row=row, column=col).value is None:
                                    device_daddr6_data[newdict['Device_Name']][i[7:]] = ""
                                else:
                                    device_daddr6_data[newdict['Device_Name']][i[7:]] = str(
                                        ws.cell(row=row, column=col).value)
                            if i[0:6] == "vpn_OL":
                                if ws.cell(row=row, column=col).value is None:
                                    device_vpn_data[newdict['Device_Name']][i[4:]] = ""
                                else:
                                    device_vpn_data[newdict['Device_Name']][i[4:]] = str(
                                        ws.cell(row=row, column=col).value)

                            col += 1

                    AllDevicesList.append(newdict)
        else:
            AllDevicesList = "worksheet"

    except Exception as e:
        AllDevicesList = "workbook"
        print(e)

    wb = None
    return AllDevicesList, headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data, device_daddr6_data, device_vpn_data


def get_workspace():
    if fmgurl.find("fortimanager.forticloud.com") != -1:
        print("### This is FortiManager Cloud - Workspacemode not supported")
        workspacemode = 0
    else:
        requestid = 1
        jsondata = {
            "method": "get",
            "params": [
                {
                    "url": "/cli/global/system/global"
                }
            ],
            "id": requestid,
            "session": fmg_sessionid
        }

        # print("Request:")
        # print(json.dumps(jsondata, indent=4, sort_keys=True))
        res = session.post(fmgurl, json=jsondata, verify=False)
        response = json.loads(res.text)
        # print("Response:")
        # print(json.dumps(response, indent=4, sort_keys=True))

        try:
            workspacemode = response['result'][0]['data']['workspace-mode']
        except:
            workspacemode = 3
    return workspacemode


def lock_adom(adom):
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvmdb/adom/" + adom + "/workspace/lock",
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))
    return response['result'][0]['status']['message']


def unlock_adom(adom):
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvmdb/adom/" + adom + "/workspace/unlock",
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))
    return response['result'][0]['status']['message']


def workspace_commit(adom):
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/devprof/adom/" + adom + "/default",
                "data": {
                    "description": str(time.time())
                }
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))



    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvmdb/adom/" + adom + "/workspace/commit",
            }
        ],
        "id": 1,
        "session": fmg_sessionid
    }
    print("Request:")
    print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    print("Response:")
    print(json.dumps(response, indent=4, sort_keys=True))
    return response['result'][0]['status']['message']


def get_meta():
    requestid = 2
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/dvmdb/_meta_fields/device",
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    meta_json = json.loads(res.text)
    return meta_json['result'][0]['data']


def create_meta(newname):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/dvmdb/_meta_fields/device",
                "data": {
                    "importance": 0,
                    "length": 255,
                    "name": newname,
                    "status": 1
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    try:
        metacreate_data = json.loads(res.text)
        ret_meta = "Meta Field " + newname + " created.<br>\n"
    except:
        ret_meta = "Failed to create Meta Field " + newname
    return ret_meta


def track_model_task(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)

    return ret_status


def add_model_device(adomname, devicename, sn, platform, prefer_img, fmg_adom_osver, fmg_adom_mr):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvm/cmd/add/device",
                "data": {
                    "adom": adomname,
                    "flags": [
                        "create_task",
                        "nonblocking"
                    ],
                    "device": {
                        "name": devicename,
                        "adm_usr": "admin",
                        "adm_pass": "",
                        "platform_str": platform,
                        "prefer_img_ver": prefer_img,
                        "mgmt_mode": 3,
                        "flags": 67371040,
                        "sn": sn,
                        "os_ver": fmg_adom_osver,
                        "mr": fmg_adom_mr
                    }
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['taskid'])
    return last_task

def add_ha_model_device(adomname, devicename, sn, platform, prefer_img, fmg_adom_osver, fmg_adom_mr, ha_sn, ha_password, ha_clustername, ha_groupid):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "dvm/cmd/add/device",
                "data": {
                    "adom": adomname,
                    "flags": [
                        "create_task",
                        "nonblocking"
                    ],
                    "device": {
                        "name": devicename,
                        "adm_usr": "admin",
                        "adm_pass": "",
                        "platform_str": platform,
                        "prefer_img_ver": prefer_img,
                        "mgmt_mode": 3,
                        "flags": 67371040,
                        "sn": sn,
                        "ha_mode": 1,
                        "ha_group_name": ha_clustername,
                        "ha_group_id": ha_groupid,
                        "ha_slave": [
                            {
                                "idx": 0,
                                "sn": sn,
                                "name": str(ha_clustername) + "-0",
                                "role": 1,
                                "prio": 255
                            },
                            {
                                "sn": ha_sn,
                                "prio": 128,
                                "idx": 1,
                                "name": str(ha_clustername) + "-1",
                                "role": 0
                            }
                        ],
                        "extra commands": [
                            {
                                "method": "update",
                                "params": [
                                    {
                                        "url": "/pm/config/device/%s/global/system/ha",
                                        "data": {
                                            "password": ha_password,
                                            "hbdev": [],
                                            "monitor": []
                                        }
                                    }
                                ]
                            }
                        ],
                        "os_ver": fmg_adom_osver,
                        "mr": fmg_adom_mr
                    }
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['taskid'])
    return last_task



def update_device(adom, devicename):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "meta fields": device_meta_data[devicename],
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)


def add_device_coords(devicename, adom, long, lat):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "longitude": long,
                    "latitude": lat
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print(res.text)
    json_devcoords = json.loads(res.text)
    status_devcoords = json_devcoords['result'][0]['status']['message']
    return status_devcoords


def change_admpass(devicename, adom, newpass):
    requestid = 1
    jsondata = {
        "method": "update",
        "params": [
            {
                "url": "/dvmdb/adom/" + adom + "/device/" + devicename,
                "data": {
                    "adm_pass": newpass
                }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("Change admin password: " + res.text)
    json_admpw = json.loads(res.text)
    status_admpw = json_admpw['result'][0]['status']['message']
    return status_admpw


def assign_cli_template(adom, template, devicename):
    ## template or template group
    template_string = "template"

    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/template-group/" + template
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_res = json.loads(res.text)
    if json_res['result'][0]['status']['message'] == "OK":
        template_string = "template-group"

    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/" + template_string + "/" + template + "/scope member",
                "data":
                    [
                        {
                            "name": devicename,
                            "vdom": "root"
                        }
                    ]

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    # print("Request:")
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_assignclitemplate = json.loads(res.text)
    # print("Response:")
    # print(json.dumps(json_assignclitemplate, indent=4, sort_keys=True))
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def unassign_cli_template(adom, template, devicename):
    ## template or template group
    template_string = "template"

    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/template-group/" + template
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_res = json.loads(res.text)
    if json_res['result'][0]['status']['message'] == "OK":
        template_string = "template-group"

    requestid = 1
    jsondata = {
        "method": "delete",
        "params": [
            {
                "url": "/pm/config/adom/" + adom + "/obj/cli/" + template_string + "/" + template + "/scope member",
                "data":
                    [
                        {
                            "name": devicename,
                            "vdom": "root"
                        }
                    ]

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_assignclitemplate = json.loads(res.text)
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def quickinstall(adom, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "/securityconsole/install/device",
                "data": {
                    "adom": adom,
                    "scope": [
                        {
                            "name": devicename,
                            "vdom": vdom
                        }
                    ]
                }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_quickinstall = json.loads(res.text)
    taskid_qi = json_quickinstall['result'][0]['data']['task']
    return taskid_qi


def track_quickinstall(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)
    return ret_status


def add_install_target(device, adomname, vdomname, pkg):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "pm/pkg/adom/" + adomname + "/" + pkg + "/scope member",
                "data": [
                    {
                        "name": device,
                        "vdom": vdomname
                    },
                ]
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_assignppkg = json.loads(res.text)
    status_ppkg = json_assignppkg['result'][0]['status']['message']
    return status_ppkg


def add_device_to_group(device, adomname, vdomname, groupname):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "dvmdb/adom/" + adomname + "/group/" + groupname + "/object member",
                "data": [
                    {
                        "name": device,
                        "vdom": vdomname
                    },
                ]
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_devgroup = json.loads(res.text)
    status_devgroup = json_devgroup['result'][0]['status']['message']
    return status_devgroup


def install_pkg(pkg, adomname, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "securityconsole/install/package",
                "data":
                    {"pkg": pkg,
                     "adom": adomname,
                     "scope": [
                         {
                             "name": devicename,
                             "vdom": vdom
                         }
                     ]
                     }
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print()
    response = json.loads(res.text)
    last_task = str(response['result'][0]['data']['task'])
    return last_task


def track_policyinstall(taskid):
    complete = 0
    while complete == 0:
        requestid = 1
        jsondata = {
            "method": "get",
            "params":
                [
                    {
                        "url": "/task/task/" + str(taskid)
                    }

                ],
            "id": requestid,
            "session": fmg_sessionid
        }
        res = session.post(fmgurl, json=jsondata, verify=False)
        task_response = json.loads(res.text)
        ret_status = False
        if task_response['result'][0]['data']['percent'] == 100:
            complete = 1
            if task_response['result'][0]['data']['num_err'] == 0:
                ret_status = True
            else:
                ret_status = False
        else:
            try:
                eel.sleep(1)
            except:
                time.sleep(1)
    return ret_status


def add_policy_interface_member(adomname, newinterfacename, realinterface, devicename):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/dynamic/interface/" + newinterfacename + "/dynamic_mapping",
                "data":
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": "root"
                            }
                        ],
                        "local-intf": realinterface,
                        "intrazone-deny": 0
                    }

            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    # print("Request:")
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_mapdint = json.loads(res.text)
    # print("Response:")
    # print(json.dumps(json_mapdint, indent=4, sort_keys=True))
    status_mapdint = json_mapdint['result'][0]['status']['message']
    return status_mapdint


def add_sdwaninterface_mapping(adomname, devicename, interfacename, vdom):
    ## get settings for base SDWAN interface
    requestid = 1

    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "/pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/" + interfacename,
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    json_sdwanint_res = json.loads(res.text)
    if json_sdwanint_res['result'][0]['status']['message'] == "OK":
        json_sdwanint = json_sdwanint_res['result'][0]['data']
        json_sdwanint.pop('dynamic_mapping', None)
        json_sdwanint.pop('obj seq', None)
        json_sdwanint.pop('name', None)

        json_sdwanint["_scope"] = [
            {
                "name": devicename,
                "vdom": vdom
            }
        ]

        proceed_makesdwanint = 0
        for key in device_sdwanint_data[devicename][interfacename]:
            proceed_makesdwanint = 1
            json_sdwanint[key] = device_sdwanint_data[devicename][interfacename][key]

        if proceed_makesdwanint == 1:
            requestid = 1
            jsondata = {
                "method": "add",
                "params": [
                    {
                        "url": "pm/config/adom/" + adomname + "/obj/dynamic/virtual-wan-link/members/" + interfacename + "/dynamic_mapping",
                        "data": json_sdwanint
                    }
                ],
                "id": requestid,
                "session": fmg_sessionid
            }
            res = session.post(fmgurl, json=jsondata, verify=False)
            print(json_sdwanint)
            print(res.text)
            json_mapsdwanint = json.loads(res.text)
            status_mapsdwanint = json_mapsdwanint['result'][0]['status']['message']
        else:
            status_mapsdwanint = "NoData"
    else:
        status_mapsdwanint = json_sdwanint_res['result'][0]['status']['message']
    return status_mapsdwanint


def assign_sdwan_template(adom, sdwantemplate, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "add",
        "params": [
            {
                "url": "/pm/wanprof/adom/" + adom + "/" + sdwantemplate + "/scope member",
                "data": [
                    {
                        "name": devicename,
                        "vdom": vdom
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_assignclitemplate = json.loads(res.text)
    status_assignclitemplate = json_assignclitemplate['result'][0]['status']['message']
    return status_assignclitemplate


def add_daddr(adomname, daddrobj, newaddr, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/firewall/address/" + daddrobj
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)

    current_int_result = json.loads(res.text)
    if current_int_result['result'][0]['status']['message'] == "OK":
        current_int = current_int_result['result'][0]['data']

        result_msg = "unknown error"

        submit = False

        if current_int['type'] == 0:
            try:
                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "allow-routing": current_int['allow-routing'],
                        "subnet": [
                            str(ipaddress.ip_network(newaddr).network_address),
                            str(ipaddress.ip_network(newaddr).netmask)
                        ],
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not decode ip address into network_address/netmask"
        elif current_int['type'] == 1:
            try:
                newaddr.strip(" ")
                splitaddr = newaddr.split("-")

                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "allow-routing": current_int['allow-routing'],
                        "end-ip": splitaddr[1].strip(),
                        "start-ip": splitaddr[0].strip()
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not calculate IP RANGE"

        if submit is True:
            requestid = 1
            jsondata = {
                "method": "add",
                "params": [
                    {
                        "url": "pm/config/adom/" + adomname + "/obj/firewall/address/" + daddrobj + "/dynamic_mapping",
                        "data": addrsettings
                    }
                ],
                "id": requestid,
                "session": fmg_sessionid
            }
            res = session.post(fmgurl, json=jsondata, verify=False)
            # print(res.text)
            json_result = json.loads(res.text)
            result_msg = json_result['result'][0]['status']['message']
    else:
        result_msg = current_int_result['result'][0]['status']['message']
    return result_msg


def add_daddr6(adomname, daddrobj, newaddr, devicename, vdom):
    requestid = 1
    jsondata = {
        "method": "get",
        "params": [
            {
                "url": "pm/config/adom/" + adomname + "/obj/firewall/address6/" + daddrobj
            }
        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)

    current_int_result = json.loads(res.text)
    if current_int_result['result'][0]['status']['message'] == "OK":
        current_int = current_int_result['result'][0]['data']

        result_msg = "unknown error"

        submit = False

        if current_int['type'] == 0:
            try:
                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "ip6": newaddr
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not decode ip address into network_address/netmask"
        elif current_int['type'] == 1:
            try:
                newaddr.strip(" ")
                splitaddr = newaddr.split("-")

                addrsettings = [
                    {
                        "_scope": [
                            {
                                "name": devicename,
                                "vdom": vdom
                            }
                        ],
                        "end-ip": splitaddr[1].strip(),
                        "start-ip": splitaddr[0].strip()
                    }
                ]
                submit = True
            except:
                result_msg = "WARNING: Could not calculate IP RANGE"

        if submit is True:
            requestid = 1
            jsondata = {
                "method": "add",
                "params": [
                    {
                        "url": "pm/config/adom/" + adomname + "/obj/firewall/address6/" + daddrobj + "/dynamic_mapping",
                        "data": addrsettings
                    }
                ],
                "id": requestid,
                "session": fmg_sessionid
            }
            res = session.post(fmgurl, json=jsondata, verify=False)
            # print(res.text)
            json_result = json.loads(res.text)
            result_msg = json_result['result'][0]['status']['message']
    else:
        result_msg = current_int_result['result'][0]['status']['message']
    return result_msg


def add_cert_template(device, adom, cert):
    requestid = 1
    jsondata = {
        "method": "exec",
        "params": [
            {
                "url": "/securityconsole/sign/certificate/template",
                "data": [
                    {
                        "adom": adom,
                        "scope": [
                            {
                                "name": device,
                                "vdom": ""
                            }
                        ],
                        "template": "adom/" + adom + "/obj/certificate/template/" + cert
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    # print(res.text)
    json_assigncerttemplate = json.loads(res.text)
    status_assigncerttemplate = json_assigncerttemplate['result'][0]['status']['message']
    return status_assigncerttemplate


def add_vpn_overlay(adom, overlayname, authpasswd):
    # Adds a VPN Community to FortiManager
    # @darryl
    requestid = 1
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/config/adom/" + adom + "/obj/vpnmgr/vpntable",
                "data": [
                    {
                        "name": overlayname,
                        "description": "Overlay Created by ZTP Tool",
                        "topology": 2,
                        "psk-auto-generate": "enable",
                        "ike1keylifesec": 28800,
                        "ike1dpd": 1,
                        "ike1natkeepalive": 10,
                        "ike2keylifesec": 1800,
                        "ike2keylifekbs": 5120,
                        "ike2keepalive": 1,
                        "intf-mode": 0,
                        "fcc-enforcement": 0,
                        "ike-version": 1,
                        "negotiate-timeout": 30,
                        "inter-vdom": 0,
                        "auto-zone-policy": 0,
                        "npu-offload": 1,
                        "authmethod": 1,
                        "ike1dhgroup": 12,
                        "dpd": 3,
                        "localid-type": 0,
                        "ike1mode": 1,
                        "ike1nattraversal": 1,
                        "ike1proposal": [
                            "aes128-sha256",
                            "aes256-sha256"
                        ],
                        "ike2autonego": 0,
                        "ike2dhgroup": 12,
                        "ike2keylifetype": 1,
                        "pfs": 1,
                        "ike2proposal": [
                            "aes128-sha256",
                            "aes256-sha256"
                        ],
                        "replay": 1
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("### add_vpn_overlay ")
    json_add_vpn_overlay = json.loads(res.text)
    status_add_vpn_overlay = json_add_vpn_overlay['result'][0]['status']['message']
    return status_add_vpn_overlay


def add_vpn_hub(adom, overlayname, interface, authpasswd, devicename, vdom, oNetwork):
    # Adds a hub to an Existing VPN community in FortiManager

    # need to add Check Overlay Exists/Check Community Exists?
    # @Darryl
    # Enhancement - Need to update Exiting Overlay\Node ID number, otherwise use a new ID.
    # Note - this currently uses ID 0 - which means next available ID number - if this imports twice you will get two entries

    oTemp = ipaddress.ip_network(oNetwork, strict=False)

    requestid = 1
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/config/adom/" + adom + "/obj/vpnmgr/node",
                "data": [
                    {
                        "id": 0,
                        "protected_subnet": {
                            "addr": "all",
                            "seq": 1
                        },
                        "scope member": {
                            "name": devicename,
                            "vdom": "root"
                        },
                        "vpntable": overlayname,
                        "role": 0,
                        "iface": interface,
                        "hub_iface": [],
                        "peer": [],
                        "automatic_routing": 0,
                        "mode-cfg": 1,
                        "mode-cfg-ip-version": 0,
                        "ipv4-start-ip": str(oTemp[10]),
                        "ipv4-end-ip": str(oTemp[-1]),
                        "ipv4-netmask": str(oTemp.netmask),
                        "net-device": 0,
                        "tunnel-search": 1,
                        "extgwip": [],
                        "extgw_hubip": [],
                        "extgw_p2_per_net": 0,
                        "route-overlap": 0,
                        "vpn-zone": [],
                        "spoke-zone": [],
                        "vpn-interface-priority": 0,
                        "auto-configuration": 1,
                        "dns-service": 5,
                        # "dhcp-server": 1,
                        "ipsec-lease-hold": 60,
                        "add-route": 0,
                        "assign-ip": 1,
                        "assign-ip-from": 0,
                        "authusrgrp": [],
                        "dns-mode": 1,
                        "exchange-interface-ip": 0,
                        # "exchange-interface-ip": 0,
                        "peergrp": [],
                        "peertype": 1,
                        "unity-support": 1,
                        "xauthtype": 1
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("### add_vpn_hub ")
    json_addvpnhub = json.loads(res.text)
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    status_addvpnhub = json_addvpnhub['result'][0]['status']['message']
    return status_addvpnhub


def add_vpn_branch(adom, overlayname, interface, authpasswd, devicename, vdom):
    # Adds a node to an Existing VPN community in FortiManager

    # need to add Check Overlay Exists/Check Community Exists?
    # @Darryl
    # Enhancement - Need to update Exiting Overlay\Node ID number, otherwise use a new ID.
    # Note - this currently uses ID 0 - which means next available ID number - if this imports twice you will get two entries

    requestid = 1
    jsondata = {
        "method": "set",
        "params": [
            {
                "url": "pm/config/adom/" + adom + "/obj/vpnmgr/node",
                "data": [
                    {
                        "protected_subnet": {
                            "addr": "all",
                            "seq": 1
                        },
                        "scope member": {
                            "name": devicename,
                            "vdom": "root"
                        },
                        "vpntable": overlayname,
                        "role": 1,
                        "usrgrp": [],
                        "iface": interface,
                        "automatic_routing": 0,
                        "extgwip": [],
                        "extgw_hubip": [],
                        "extgw_p2_per_net": 0,
                        "route-overlap": 0,
                        "vpn-zone": [],
                        "spoke-zone": [],
                        "vpn-interface-priority": 0,
                        "auto-configuration": 1,
                        "ipsec-lease-hold": 60,
                        "add-route": 0,
                        "assign-ip": 0,
                        "assign-ip-from": 0,
                        "exchange-interface-ip": 1,
                        "mode-cfg": 1,
                        "mode-cfg-ip-version": 0,
                        "net-device": 1,
                        "peergrp": [],
                        "peertype": 8,
                        "tunnel-search": 0,
                        "unity-support": 1,
                        "xauthtype": 1
                    }
                ]
            }

        ],
        "id": requestid,
        "session": fmg_sessionid
    }
    res = session.post(fmgurl, json=jsondata, verify=False)
    print("### add_vpn_branch")
    json_addvpnnode = json.loads(res.text)
    # print(json.dumps(jsondata, indent=4, sort_keys=True))
    status_addvpnnode = json_addvpnnode['result'][0]['status']['message']
    return status_addvpnnode


##====================================================================================
##====================================================================================
##====================================================================================
##====================================================================================
##====================================================================================


@eel.expose
def btn_checkxlsx(filename, fmghost, fmguser, fmgpasswd, fmgadom):
    global fmg_user
    global fmg_passwd
    global fmgurl
    global fmg_adom
    global fmg_sessionid
    global device_meta_data
    global device_dint_data
    global device_sdwanint_data
    global device_vpn_data
    global alldevices

    qi_status = False
    fmg_adom = fmgadom
    fmg_user = fmguser
    fmg_passwd = fmgpasswd
    fmgurl = "https://" + fmghost + "/jsonrpc"
    fmg_sessionid = None

    return_html = ""
    sendupdate(return_html)

    alldevices, headings, device_meta_data, device_dint_data, device_sdwanint_data, device_daddr_data, device_daddr6_data, device_vpn_data = openbook(
        filename)

    if alldevices == "workbook":
        return_html += "Load Excel workbook failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
    elif alldevices == "worksheet":
        return_html += "Load Excel worksheet failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
        return_html += "Device_Name not found in cell A1 of worksheet named 'Devices', 'Sheet1', or the active worksheet.<br>\n"
    else:
        return_html += "Load Excel workbook successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
        ##login to FMG

        sendupdate(return_html)

        requestid = 1

        jsondata = {'method': 'exec',
                    'params': [{'url': '/sys/login/user', 'data': {'user': fmg_user, 'passwd': fmg_passwd}}],
                    'id': requestid}

        try:
            res = session.post(fmgurl, json=jsondata, verify=False, timeout=4)
            try:
                login_data = json.loads(res.text)
                fmg_sessionid = login_data['session']
                return_html += "FortiManager login successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
            except:
                return_html += "FortiManager login failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
        except requests.exceptions.RequestException:
            return_html += "FortiManager connection failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    if fmg_sessionid is not None:
        proceed = True
    else:
        proceed = False

    sendupdate(return_html)
    ### validity checks
    if proceed == True:
        print("### validity checks")
        ## Does ADOM exist in FMG

        jsondata = {"method": "get", "params": [{"url": "dvmdb/adom/" + fmg_adom}], "id": requestid,
                    "session": fmg_sessionid}
        res = session.post(fmgurl, json=jsondata, verify=False)
        json_result = json.loads(res.text)
        print("-- ## Does ADOM exist in FMG v1 -- ")
        # print(json_result['result'][0]['status']['message'])
        if json_result['result'][0]['status']['message'] != "OK":
            return_html += "FortiManager ADOM does not exist <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
            proceed = False
        else:
            ## Record ADOM FGT Version
            fmg_adom_osver = json_result['result'][0]['data']['os_ver']
            fmg_adom_mr = json_result['result'][0]['data']['mr']

        ## Get workspace mode

        workspacemode = get_workspace()
        if workspacemode == 3:
            return_html += "Error determing workspace mode <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
            proceed = False
        elif workspacemode == 2:
            return_html += "FortiManager is in workflow mode (not supported) <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
            proceed = False
        elif workspacemode == 1:
            return_html += "FortiManager is in workspace mode <span class=\"glyphicon glyphicon-info-sign\" style=\"color:blue\"></span><br>\n"
            # proceed = False
        elif workspacemode == 0:
            return_html += "FortiManager has workspace mode disabled <span class=\"glyphicon glyphicon-info-sign\" style=\"color:blue\"></span><br>\n"

        ## Lock ADOM is workspace mode is enabled

        if workspacemode == 1:
            get_lock = lock_adom(fmg_adom)
            if get_lock == "OK":
                return_html += "Lock ADOM \"" + fmg_adom + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
            else:
                return_html += "Lock ADOM \"" + fmg_adom + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                proceed = False
        ## check for required fields in headings
        required_headings = ["Device_Name", "Platform", "Device_SN", "CLI_Template", "Post_CLI_Template",
                             "Policy_Package",
                             "SDWAN_Template"]
        for req_heading in required_headings:
            if req_heading not in headings:
                return_html += "Excel File does not have required heading \"" + req_heading + "\" <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                proceed = False

    ### Check for Meta Fields and Create if they dont exist
    if proceed == True:
        print("### Create/Check for Meta Fields")

        metafields = get_meta()
        for field in headings:
            if field[0:5] == "meta_":
                try:
                    IPDICT = next(item for item in metafields if item["name"] == field[5:])
                except:
                    create_meta(field[5:])

        ## check for automatic Meta field for Device_Name and Device_SN
        try:
            IPDICT = next(item for item in metafields if item["name"] == "Device_Name")
        except:
            create_meta("Device_Name")
        try:
            IPDICT = next(item for item in metafields if item["name"] == "Device_SN")
        except:
            create_meta("Device_SN")

    ### Create Model Devices
    if proceed == True:
        print("### Create Model Devices")
        sendupdate(return_html)

        for devicedata in alldevices:

            try:
                devicemode = devicedata['ztpmode']
            except:
                devicemode = "default"

            print(devicemode)


            if devicemode == "updatemeta":
                ### Add meta data to device
                print("   ### Add meta data to device")
                update_device(fmg_adom, devicedata['Device_Name'])
                return_html += "Update meta data for (" + devicedata['Device_Name'] +") successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
            else:
                if "HA_SN" in devicedata:
                    print("   ### Create HA Model Device " + devicedata['Device_Name'])
                    return_html += "<br>\n <b> >> Adding HA Device [ " + devicedata['Device_Name'] + " ] </b><br>\n"
                    add_dev_status = track_model_task(
                        add_ha_model_device(fmg_adom, devicedata['Device_Name'], devicedata['Device_SN'],
                                         devicedata['Platform'], devicedata['Upgrade_Ver'], fmg_adom_osver,
                                         fmg_adom_mr, devicedata['HA_SN'], devicedata["HA_Password"], devicedata['HA_ClusterName'], devicedata['HA_GroupID']))

                    sendupdate(return_html)
                else:
                    print("   ### Create Model Device " + devicedata['Device_Name'])
                    return_html += "<br>\n <b> >> Adding Device [ " + devicedata['Device_Name'] + " ] </b><br>\n"
                    add_dev_status = track_model_task(
                        add_model_device(fmg_adom, devicedata['Device_Name'], devicedata['Device_SN'],
                                         devicedata['Platform'], devicedata['Upgrade_Ver'], fmg_adom_osver, fmg_adom_mr))

                    sendupdate(return_html)

                if add_dev_status == True:
                    return_html += "Adding model device (" + devicedata['Device_Name'] + "/" + devicedata[
                        'Device_SN'] + ") successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                else:
                    return_html += "Adding model device (" + devicedata['Device_Name'] + "/" + devicedata[
                        'Device_SN'] + ") failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                if add_dev_status == True:
                    ## Add device to device group
                    print("   ### Add device to device group")
                    if "Device_Group" in devicedata:
                        if devicedata['Device_Group'] == "" or devicedata['Device_Group'] is None:
                            return_html += "Assign Device Group {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_devgroup = add_device_to_group(devicedata['Device_Name'], fmg_adom, 'root',
                                                                  devicedata['Device_Group'])

                            if status_devgroup == "OK":
                                return_html += "Assign Device Group \"" + devicedata[
                                    'Device_Group'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Assign Device Group \"" + devicedata[
                                    'Device_Group'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    ### Add coordinates to device
                    print("   ### Add coordinates to device")
                    if "Device_Longitute" in devicedata and "Device_Latitute" in devicedata:
                        if devicedata['Device_Longitute'] == "" or devicedata['Device_Longitute'] is None or devicedata[
                            'Device_Latitute'] == "" or devicedata['Device_Latitute'] is None:
                            return_html += "Assign Device Coordinates {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_addcoords = add_device_coords(devicedata['Device_Name'], fmg_adom,
                                                                 devicedata['Device_Longitute'],
                                                                 devicedata['Device_Latitute'])

                            if status_addcoords == "OK":
                                return_html += "Assign Device Coordinates successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Assign Device Coordinates failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    ## Change device password for admin user if Device_Adminpassword exists in excel sheet
                    print("   ### Change device password")
                    if "Device_Adminpassword" in devicedata:
                        if devicedata['Device_Adminpassword'] == "" or devicedata['Device_Adminpassword'] is None:
                            return_html += "Change Device Admin Password {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_changeadmpass = change_admpass(devicedata['Device_Name'], fmg_adom,
                                                                  devicedata['Device_Adminpassword'])

                            if status_changeadmpass == "OK":
                                return_html += "Change Device Admin Password successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Change Device Admin Password failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"


                    ### Add certificate template to device
                    print("   ### Add certificate template to device")
                    if "Cert_Template" in devicedata:
                        if devicedata['Cert_Template'] == "" or devicedata['Cert_Template'] is None:
                            return_html += "Assign Certificate Template {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_certtemplate = add_cert_template(devicedata['Device_Name'], fmg_adom,
                                                                 devicedata['Cert_Template'])

                            if status_certtemplate == "OK":
                                return_html += "Assign Certificate Template successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Assign Certificate Template failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"


                    ### Add meta data to device
                    print("   ### Add meta data to device")
                    update_device(fmg_adom, devicedata['Device_Name'])

                    ### Assign Initial CLI Template
                    print("   ### Assign Initial CLI Template")
                    status_clitemp = ""
                    qi_status = False
                    if devicedata['CLI_Template'] == "" or devicedata['CLI_Template'] is None:
                        return_html += "Assign CLI Template {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        qi_status = True
                    else:
                        status_clitempgrp = assign_cli_template(fmg_adom, devicedata['CLI_Template'],
                                                                devicedata['Device_Name'])
                        if status_clitempgrp == "OK":
                            return_html += "Assign CLI Template \"" + devicedata[
                                'CLI_Template'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            status_clitemp = "OK"
                        else:
                            return_html += "Assign CLI Template \"" + devicedata[
                                'CLI_Template'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    if status_clitemp == "OK":
                        if workspacemode == 1:
                            workspace_commit(fmg_adom)


                        ##Install Device Settings
                        qi_status = track_quickinstall(quickinstall(fmg_adom, devicedata['Device_Name'], 'root'))
                        if qi_status:
                            return_html += "Quick install device settings successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                        else:
                            return_html += "Quick install device settings failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                        ##Unassign CLI Template
                        unassign_cli_template(fmg_adom, devicedata['CLI_Template'], devicedata['Device_Name'])

                        if workspacemode == 1:
                            workspace_commit(fmg_adom)

                if qi_status == True:

                    ## Quick Install Sucessful, assign policy package etc
                    sendupdate(return_html)
                    ##map interfaces

                    for key in device_dint_data[devicedata['Device_Name']]:
                        if device_dint_data[devicedata['Device_Name']][key] == "":
                            return_html += "Add dynamic map for interface \"" + key + "\" {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_mapdint = add_policy_interface_member(fmg_adom, key,
                                                                         device_dint_data[devicedata['Device_Name']][key],
                                                                         devicedata['Device_Name'])

                            if status_mapdint == "OK":
                                return_html += "Add dynamic map for interface \"" + key + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Add dynamic map for interface \"" + key + "\" failed ><span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                                return_html += status_mapdint + "<br>\n"

                    ### MAP DYNAMIC ADDRESS OJBECTS
                    ## ipv4
                    for key in device_daddr_data[devicedata['Device_Name']]:
                        if device_daddr_data[devicedata['Device_Name']][key] == "":
                            return_html += "Add dynamic map for address \"" + key + "\" {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_mapdaddr = add_daddr(fmg_adom, key, device_daddr_data[devicedata['Device_Name']][key],
                                                        devicedata['Device_Name'], 'root')
                            if status_mapdaddr == "OK":
                                return_html += "Add dynamic map for address \"" + key + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Add dynamic map for address \"" + key + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                                return_html += status_mapdaddr + "<br>\n"

                    ### Add Branch to Central VPN Manager (Darryl)

                    for key in device_vpn_data[devicedata['Device_Name']]:  ## e.g. key = vpn_OL_INET, vpn_OL_MPLS, ishub
                        if device_vpn_data[devicedata['Device_Name']][key] == "":
                            return_html += "Add vpn node for device \"" + key + "\" {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            # key = the overlay name
                            add_vpn_overlay(fmg_adom, key, "")

                            print("     Is this device a vpn hub: " + devicedata['vpn_IsHub'] + " / Overlayname: " + key)
                            if devicedata['vpn_IsHub'] in ["true", "yes", "hub", "1"]:
                                print('    vpn_Subnet_' + key + "=" + devicedata['vpn_Subnet_' + key])
                                status_mapvpnnode = add_vpn_hub(fmg_adom, key,
                                                                device_vpn_data[devicedata['Device_Name']][key],
                                                                "", devicedata['Device_Name'], fmg_adom,
                                                                devicedata['vpn_Subnet_' + key])
                            else:
                                status_mapvpnnode = add_vpn_branch(fmg_adom, key,
                                                                   device_vpn_data[devicedata['Device_Name']][key],
                                                                   "", devicedata['Device_Name'], fmg_adom)

                            if status_mapvpnnode == "OK":
                                return_html += "Add vpnnode map for device \"" + key + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Add vpnnode map for device \"" + key + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                                return_html += status_mapdaddr + "<br>\n"

                    ## ipv6
                    for key in device_daddr6_data[devicedata['Device_Name']]:
                        if device_daddr6_data[devicedata['Device_Name']][key] == "":
                            return_html += "Add dynamic map for address6 \"" + key + "\" {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            status_mapdaddr6 = add_daddr6(fmg_adom, key, device_daddr6_data[devicedata['Device_Name']][key],
                                                          devicedata['Device_Name'], 'root')
                            if status_mapdaddr6 == "OK":
                                return_html += "Add dynamic map for address6 \"" + key + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                            else:
                                return_html += "Add dynamic map for address6 \"" + key + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                                return_html += status_mapdaddr6 + "<br>\n"

                    ### MAP SDWAN Interfaces
                    for key in device_sdwanint_data[devicedata['Device_Name']]:
                        status_mapsdwanint = add_sdwaninterface_mapping(fmg_adom, devicedata['Device_Name'], key, 'root')
                        if status_mapsdwanint == "OK":
                            return_html += "Add dynamic SDWAN Map for interface \"" + key + "\" succcessful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                        elif status_mapsdwanint == "NoData":
                            return_html += "Add SD-WAN interface map for \"" + key + "\" {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                        else:
                            return_html += "Add dynamic SDWAN Map for interface \"" + key + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                            return_html += status_mapsdwanint + "<br>\n"

                    ## Assign SDWAN Template
                    if devicedata['SDWAN_Template'] == "" or devicedata['SDWAN_Template'] is None:
                        return_html += "Assign SDWAN template \"{not defined}\" <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                    else:
                        status_assignsdwantemplate = assign_sdwan_template(fmg_adom, devicedata['SDWAN_Template'],
                                                                           devicedata['Device_Name'], 'root')
                        if status_assignsdwantemplate == "OK":
                            return_html += "Assign SDWAN template \"" + devicedata[
                                'SDWAN_Template'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                        else:
                            return_html += "Assign SDWAN template \"" + devicedata[
                                'SDWAN_Template'] + "\" failed ><span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    ## Add install Target
                    status_add_inst_trgt = add_install_target(devicedata['Device_Name'], fmg_adom, "root",
                                                              devicedata['Policy_Package'])
                    if status_add_inst_trgt == "OK":
                        return_html += "Assign policy package \"" + devicedata[
                            'Policy_Package'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                    else:
                        return_html += "Assign policy package \"" + devicedata[
                            'Policy_Package'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    ## Assign Post CLI Template
                    if devicedata['Post_CLI_Template'] == "" or devicedata['Post_CLI_Template'] is None:
                        return_html += "Assign Post CLI Template {not defined} <span class=\"glyphicon glyphicon-info-sign\" style=\"color:orange\"></span><br>\n"
                    else:
                        status_clitempgrp = assign_cli_template(fmg_adom, devicedata['Post_CLI_Template'],
                                                                devicedata['Device_Name'])
                        if status_clitempgrp == "OK":
                            return_html += "Assign Post CLI Template \"" + devicedata[
                                'Post_CLI_Template'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                        else:
                            return_html += "Assign Post CLI Template \"" + devicedata[
                                'Post_CLI_Template'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                    if status_add_inst_trgt == "OK":
                        ## install package

                        if workspacemode == 1:
                            workspace_commit(fmg_adom)

                        pkg_status = track_policyinstall(
                            install_pkg(devicedata['Policy_Package'], fmg_adom, devicedata['Device_Name'], 'root'))
                        if pkg_status == True:
                            return_html += "Install policy package \"" + devicedata[
                                'Policy_Package'] + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                        else:
                            return_html += "Install policy package \"" + devicedata[
                                'Policy_Package'] + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

                        if pkg_status == True:
                            pass

                sendupdate(return_html)

        if workspacemode == 1:
            get_unlock = unlock_adom(fmg_adom)
            if get_unlock == "OK":
                return_html += "Unlock ADOM \"" + fmg_adom + "\" successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
            else:
                return_html += "Unlock ADOM \"" + fmg_adom + "\" failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    ### LOGOUT OF FMG
    if fmg_sessionid is not None:
        print("### LOGOUT OF FMG")
        requestid = 1
        jsondata = {'method': 'exec', 'params': [{'url': '/sys/logout'}], 'session': fmg_sessionid, 'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)

    return_html += "<br>\n<b> >> Complete! <br>\n"
    return_html += "<br>\n <a href=\"ztptool.html\">Return</a> <br>\n"

    sendupdate(return_html)


### End copy from draft

@eel.expose
def btn_checkadom(filename, fmghost, fmguser, fmgpasswd, fmgadom, fmgadomdesc):
    global fmg_user
    global fmg_passwd
    global fmgurl
    global fmg_adom
    global fmg_adomdesc
    global fmg_sessionid

    fmg_adom = fmgadom
    fmg_user = fmguser
    fmg_passwd = fmgpasswd
    fmgurl = "https://" + fmghost + "/jsonrpc"
    fmg_sessionid = None

    return_html = ""
    sendupdate(return_html)
    new_json = ""
    with open(filename) as json_data_file:
        vars = {"$(adom_name)": fmgadom, "$(adom_desc)": fmgadomdesc}

        for line in json_data_file:
            m = re.search('\$\((.+)\)', line)
            if m:
                print("Old: " + line.rstrip())
                for key in vars.keys():
                    line = line.replace(key, vars[key])

                print("New: " + line.rstrip())

            new_json += line

    try:
        adom_json = json.loads(new_json)
        return_html += "Load JSON file successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
    except:
        return_html += "Load Excel workbook failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    sendupdate(return_html)

    if adom_json:
        ##login to FMG

        requestid = 1

        jsondata = {'method': 'exec',
                    'params': [{'url': '/sys/login/user', 'data': {'user': fmg_user, 'passwd': fmg_passwd}}],
                    'id': requestid}

        try:
            res = session.post(fmgurl, json=jsondata, verify=False, timeout=4)
            try:
                login_data = json.loads(res.text)
                fmg_sessionid = login_data['session']
                return_html += "FortiManager login successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
            except:
                return_html += "FortiManager login failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
        except requests.exceptions.RequestException:
            return_html += "FortiManager connection failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    sendupdate(return_html)

    if fmg_sessionid is not None:
        for item in adom_json["settings"]:
            for dataset in item["data"]:

                jsondata = {
                    "method": item["method"],
                    "params": [
                        {
                            "url": item["url"],
                            "data": dataset

                        }
                    ],
                    "id": requestid,
                    "session": fmg_sessionid
                }
                print("Request:")
                print(json.dumps(jsondata, indent=4, sort_keys=True))
                res = session.post(fmgurl, json=jsondata, verify=False)
                response = json.loads(res.text)
                print("Response:")
                print(json.dumps(response, indent=4, sort_keys=True))
                response_url = response['result'][0]['url']
                try:
                    response_name = response['result'][0]['data']['name']
                except:
                    try:
                        response_name = jsondata['params'][0]['data']['name']
                    except:
                        response_name = ""
                if response['result'][0]['status']['message'] == "OK":
                    return_html += response_url + response_name + " <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br>\n"
                else:
                    return_html += response_url + response_name + " <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
                    return_html += " >> " + response['result'][0]['status']['message'] + "<br>\n"

                sendupdate(return_html)

    ### LOGOUT OF FMG
    if fmg_sessionid is not None:
        requestid = 1
        jsondata = {'method': 'exec', 'params': [{'url': '/sys/logout'}], 'session': fmg_sessionid, 'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)

    return_html += "<br>\n<b> >> Complete! <br>\n"
    return_html += "<br>\n <a href=\"ztptool.html\">Return</a> <br><br><br><br>&nbsp;\n"

    sendupdate(return_html)


@eel.expose
def btn_checkexportadom(fmghost, fmguser, fmgpasswd, fmgadom):
    global fmg_user
    global fmg_passwd
    global fmgurl
    global fmg_adom
    global fmg_sessionid
    global json_export
    global requestid

    fmg_adom = fmgadom
    fmg_user = fmguser
    fmg_passwd = fmgpasswd
    fmgurl = "https://" + fmghost + "/jsonrpc"
    fmg_sessionid = None
    proceed = True
    return_html = ""
    sendupdate(return_html)
    requestid = 1

    ##login to FMG

    requestid = 1

    jsondata = {'method': 'exec',
                'params': [{'url': '/sys/login/user', 'data': {'user': fmg_user, 'passwd': fmg_passwd}}],
                'id': requestid}

    try:
        res = session.post(fmgurl, json=jsondata, verify=False, timeout=4)
        try:
            login_data = json.loads(res.text)
            fmg_sessionid = login_data['session']
            return_html += "FortiManager login successful <span class=\"glyphicon glyphicon-ok\" style=\"color:green\"></span><br><br>\n"
        except:
            return_html += "FortiManager login failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"
    except requests.exceptions.RequestException:
        return_html += "FortiManager connection failed <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br>\n"

    sendupdate(return_html)

    ### validity checks
    if fmg_sessionid:
        ## Does ADOM exist in FMG

        jsondata = {"method": "get", "params": [{"url": "dvmdb/adom/" + fmg_adom}], "id": requestid,
                    "session": fmg_sessionid}
        res = session.post(fmgurl, json=jsondata, verify=False)
        json_result = json.loads(res.text)
        print("-- ## Does ADOM exist in FMG v2 -- ")
        print(json_result['result'][0]['status']['message'])
        if json_result['result'][0]['status']['message'] != "OK":
            return_html += "FortiManager ADOM does not exist <span class=\"glyphicon glyphicon-remove\" style=\"color:red\"></span><br><br/>\n"
            proceed = False

        if proceed == True:
            json_export = export_adom(fmg_adom)

            return_html += "<textarea readonly rows=\"15\" id=\"jsonexport\" class=\"form-control\" style=\"min-width: 100%\">" + json_export + "</textarea><br/>\n"

            return_html += "<div form-group><button type=\"button\" onclick=\"eel.btn_saveadom()\" class=\"btn btn-secondary\">Save As</button>&nbsp;<span id=\"filepath\"></span><br/><br/></div>"

    sendupdate(return_html)

    ### LOGOUT OF FMG
    if fmg_sessionid is not None:
        requestid = 1
        jsondata = {'method': 'exec', 'params': [{'url': '/sys/logout'}], 'session': fmg_sessionid, 'id': requestid}
        res = session.post(fmgurl, json=jsondata, verify=False)

    return_html += "<br>\n <a href=\"ztptool.html\">Return</a> <br><br><br><br>&nbsp;\n"

    sendupdate(return_html)


@eel.expose
def btn_getxlsxfile():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("XLSX Files", "*.xlsx"), ("all files", "*.*")))
    root.update()  # to make dialog close on MacOS
    print(filename)

    return filename


@eel.expose
def btn_saveadom():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    filename = filedialog.asksaveasfilename(initialdir="/", defaultextension=".json", title="Select file",
                                            filetypes=(("JSON Files", "*.json"), ("all files", "*.*")))
    root.update()  # to make dialog close on MacOS

    print(filename)

    try:
        f = open(filename, "a")
        f.write(json_export)
        f.close()
        eel.saveupdate("Saved as: " + filename)
    except:
        eel.saveupdate("Could not save as: + " + filename)


@eel.expose
def btn_getjsonfile():
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                          filetypes=(("JSON Files", "*.json"), ("all files", "*.*")))
    root.update()  # to make dialog close on MacOS
    return filename


@eel.expose
def savesettings(save_fmg, save_user, save_adom, save_path, save_pw):
    settingsfiledata = '''{
  "fmg": "%s",
  "user": "%s",
  "passwd": "%s",
  "adom": "%s",
  "path": "%s"
}
''' % (save_fmg, save_user, save_pw, save_adom, save_path)

    try:
        setting_file = open("settings.json", "wt")
        setting_file.write(settingsfiledata)
        setting_file.close()
        return ["Settings Saved", "success"]
    except:
        return ["Error: Could not save settings", "danger"]


@eel.expose
def getsettings_adom():
    try:
        with open('settings.json') as json_settings:
            settings = json.load(json_settings)
            try:
                default_fmg = settings['fmg']
            except:
                default_fmg = ""
            try:
                default_user = settings['user']
            except:
                default_user = ""
            try:
                default_passwd = settings['passwd']
            except:
                default_passwd = ""
            try:
                default_adom = settings['adom']
            except:
                default_adom = ""
        json_settings.close()
    except:
        default_fmg = ""
        default_user = ""
        default_passwd = ""
        default_adom = ""

    return_html = '''
            <div class="starter-template">
      <h4>Import ADOM</h4>
    </div>
    <div>
      <form autocomplete="off">

        <label for="fmgip">FortiManager URL</label>



        <div class="input-group mb-3">
          <div class="input-group-prepend">
            <span class="input-group-text" id="https-addon">https://</span>
          </div>
          <input type="text" class="form-control" id="fmgip" aria-describedby="https-addon" value="%s">
        </div>


        <div class="form-group">
          <label for="fmgusername">FortiManager Username</label>
          <input type="text" class="form-control" id="fmgusername" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgpassword">FortiManager Password</label>
          <input type="password" class="form-control" id="fmgpassword">
        </div>
        <div class="form-group">
          <label for="fmgadom">New FortiManager ADOM</label>
          <input type="text" class="form-control" id="fmgadom" value="">
        </div>
        <div class="form-group">
          <label for="fmgadom">New FortiManager ADOM Description</label>
          <input type="text" class="form-control" id="fmgadomdesc" value="">
        </div>        
        <div form-group>
          <button type="button" onclick="getFileADOM()" class="btn btn-secondary btn-sm">Select File</button>
          JSON Path: <span id="filepath">/</span><br/><br/>
        </div>
        <div class="form-group">
          <button type="button" onclick="processadom(document.getElementById('filepath').innerHTML)" class="btn btn-primary">Submit</button>
        </div>
      </form>

    </div>
          ''' % (default_fmg, default_user)

    eel.pageupdate(return_html)


@eel.expose
def getsettings_exportadom():
    try:
        with open('settings.json') as json_settings:
            settings = json.load(json_settings)
            try:
                default_fmg = settings['fmg']
            except:
                default_fmg = ""
            try:
                default_user = settings['user']
            except:
                default_user = ""
            try:
                default_passwd = settings['passwd']
            except:
                default_passwd = ""
            try:
                default_adom = settings['adom']
            except:
                default_adom = ""
        json_settings.close()
    except:
        default_fmg = ""
        default_user = ""
        default_passwd = ""
        default_adom = ""

    return_html = '''
            <div class="starter-template">
      <h4>Export ADOM</h4>
    </div>
    <div>
      <form autocomplete="off">

        <label for="fmgip">FortiManager URL</label>



        <div class="input-group mb-3">
          <div class="input-group-prepend">
            <span class="input-group-text" id="https-addon">https://</span>
          </div>
          <input type="text" class="form-control" id="fmgip" aria-describedby="https-addon" value="%s">
        </div>


        <div class="form-group">
          <label for="fmgusername">FortiManager Username</label>
          <input type="text" class="form-control" id="fmgusername" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgpassword">FortiManager Password</label>
          <input type="password" class="form-control" id="fmgpassword" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgadom">FortiManager ADOM to export</label>
          <input type="text" class="form-control" id="fmgadom" value="">
        </div>


        <div class="form-group">
          <button type="button" onclick="processexportadom()" class="btn btn-primary">Submit</button>
        </div>
      </form>

    </div>
          ''' % (default_fmg, default_user, default_passwd)

    eel.pageupdate(return_html)


@eel.expose
def getsettings_devices():
    try:
        with open('settings.json') as json_settings:
            settings = json.load(json_settings)
            try:
                default_fmg = settings['fmg']
            except:
                default_fmg = ""
            try:
                default_user = settings['user']
            except:
                default_user = ""
            try:
                default_passwd = settings['passwd']
            except:
                default_passwd = ""
            try:
                default_adom = settings['adom']
            except:
                default_adom = ""
            try:
                default_path = settings['path']
            except:
                default_path = "/"
        json_settings.close()
    except:
        default_fmg = ""
        default_user = ""
        default_passwd = ""
        default_adom = ""

    return_html = '''
            <div class="starter-template">
      <h4>Import Devices</h4>
    </div>
    <div>
      <form autocomplete="off">

        <label for="fmgip">FortiManager URL</label>



        <div class="input-group mb-3">
          <div class="input-group-prepend">
            <span class="input-group-text" id="https-addon">https://</span>
          </div>
          <input type="text" class="form-control" id="fmgip" aria-describedby="https-addon" value="%s">
        </div>


        <div class="form-group">
          <label for="fmgusername">FortiManager Username</label>
          <input type="text" class="form-control" id="fmgusername" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgpassword">FortiManager Password</label>
          <input type="password" class="form-control" id="fmgpassword" value="%s">
        </div>
        <div class="form-group">
          <label for="fmgadom">FortiManager ADOM</label>
          <input type="text" class="form-control" id="fmgadom" value="%s">
        </div>
        <div form-group>
          <button type="button" onclick="getFolder()" class="btn btn-secondary btn-sm">Select File</button>
          Excel Path: <span id="filepath">%s</span> <div class="float-right">
          <button type="button" class="btn btn-info btn-sm" data-toggle="modal" data-target="#savesettingsModal">Save Settings <span class="glyphicon glyphicon-floppy-save"></span></button></div><br/><br/>
        </div>
        <div class="form-group">
          <button type="button" onclick="processxlsx(document.getElementById('filepath').innerHTML)" class="btn btn-primary">Submit</button>
        </div>
      </form>

    </div>
          ''' % (default_fmg, default_user, default_passwd, default_adom, default_path)

    eel.pageupdate(return_html)


session = requests.session()

use_mode = ''
try:
    with open('browser.json') as json_browsersettings:
        browsersettings = json.load(json_browsersettings)
        use_mode = browsersettings['mode']
        use_cmdline_args = browsersettings['cmdline_args']
except:
    pass

if use_mode == "":
    try:
        eel.start('ztptool.html', size=(790, 850), disable_cache=True)
    except EnvironmentError:
        # If Chrome isn't found, fallback to Microsoft Edge on Win10 or greater
        if sys.platform in ['win32', 'win64'] and int(platform.release()) >= 10:
            eel.start('ztptool.html', size=(790, 850), disable_cache=True, mode='edge')
        else:
            raise

else:
    eel.start('ztptool.html', size=(790, 850), disable_cache=True, mode=use_mode, cmdline_args=use_cmdline_args)
